"""数据文件加载模块 — 支持 Excel (.xlsx) 和 CSV (.csv)"""

import csv
import os
from typing import Any

from openpyxl import load_workbook


def load_data(filepath: str) -> list[dict[str, Any]]:
    """加载数据文件，自动检测格式，返回 list[dict]。
    第一行为列头，后续每行一条记录。None 值转为空字符串。
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsx":
        return _load_excel(filepath)
    elif ext == ".csv":
        return _load_csv(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx 或 .csv 文件")


def get_columns(filepath: str, header_row_num: int = 1) -> list[str]:
    """快速读取列头（指定行）"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsx":
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))]
        wb.close()
        return [str(h) if h is not None else "" for h in headers]
    elif ext == ".csv":
        with open(filepath, "r", encoding=_detect_csv_encoding(filepath)) as f:
            reader = csv.reader(f)
            for _ in range(header_row_num - 1):
                next(reader, None)
            return [str(h) for h in next(reader)]
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def load_appendix_data(filepath: str, sheet_name: str | None = None) -> tuple[list[str], list[list[Any]]]:
    """加载附录 Excel 文件，返回 (列头列表, 数据行列表)。
    自动跳过前导空行，找到第一个包含实际列头的行作为表头。
    数值按 Excel 单元格格式化输出（如 0.06→6%, 4051→4,051.00）。
    """
    wb = load_workbook(filepath, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers: list[str] = []
    data: list[list[Any]] = []
    header_found = False

    for row in ws.iter_rows():
        values = [cell.value for cell in row]

        # 跳过全空行
        non_none = [v for v in values if v is not None and str(v).strip()]
        if not non_none:
            continue

        # 找表头行
        if not header_found:
            str_row = [str(v) if v is not None else "" for v in values]
            non_empty_count = sum(1 for s in str_row if s.strip())
            if non_empty_count >= len(str_row) * 0.5:
                headers = str_row
                header_found = True
            continue

        # 数据行：按 Excel 格式化每个单元格
        formatted_row = [_format_cell(cell) for cell in row]
        data.append(formatted_row)

    wb.close()
    return headers, data


def _format_cell(cell) -> Any:
    """根据 Excel 单元格的数字格式，将值格式化为字符串。"""
    val = cell.value
    if val is None:
        return None

    fmt = (cell.number_format or "").strip()

    if isinstance(val, (int, float)):
        if "%" in fmt:
            # 百分比格式：0.06 → 6%
            return f"{val * 100:.0f}%"
        elif "," in fmt or fmt in ("#,##0", "#,##0.00", "0.00", "#,##0.00_);(#,##0.00)"):
            # 千分位数字格式
            if "." in fmt:
                return f"{val:,.2f}"
            else:
                return f"{val:,.0f}"
        elif fmt == "0" or fmt == "General":
            # 整数或通用格式
            if isinstance(val, float) and val == int(val):
                return int(val)
            return val

    return val


def _load_excel(filepath: str) -> list[dict[str, Any]]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        record = {}
        for h, v in zip(headers, row):
            record[h] = v if v is not None else ""
        result.append(record)
    wb.close()
    return result


def _detect_csv_encoding(filepath: str) -> str:
    """尝试检测 CSV 文件编码"""
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                f.read(1024)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def _load_csv(filepath: str) -> list[dict[str, Any]]:
    encoding = _detect_csv_encoding(filepath)
    with open(filepath, "r", encoding=encoding) as f:
        reader = csv.DictReader(f)
        result = []
        for row in reader:
            record = {k: (v if v is not None else "") for k, v in row.items()}
            result.append(record)
    return result
