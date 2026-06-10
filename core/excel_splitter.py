"""Excel 拆分模块 — 按指定列分组拆分 Excel 文件"""

import os
import re
from typing import Any, Callable

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def _read_header_area_and_widths(
    ws, header_row_num: int
) -> tuple[list[list[Any]], dict[str, float], list[float | None]]:
    """读取表头区域（第1行到 header_row_num 行）、列宽和各行行高。

    返回:
        header_rows: 表头区域所有行的数据（每行为 list）
        column_widths: 列宽映射
        header_row_heights: 各表头行的行高列表
    """
    header_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=header_row_num, values_only=True):
        header_rows.append(list(row))

    # 列数取表头区域最大列数
    max_cols = max(len(r) for r in header_rows) if header_rows else 0

    column_widths: dict[str, float] = {}
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws.column_dimensions:
            column_widths[col_letter] = ws.column_dimensions[col_letter].width

    header_row_heights: list[float | None] = []
    for r in range(1, header_row_num + 1):
        header_row_heights.append(
            ws.row_dimensions[r].height if r in ws.row_dimensions else None
        )

    return header_rows, column_widths, header_row_heights


def _save_workbook(
    new_wb, new_ws, ws_title, header_rows, rows_data,
    column_widths, header_row_heights, out_path
):
    """写入表头区域和数据行，复制格式并保存。"""
    new_ws.title = ws_title if ws_title else "Sheet1"

    # 写入表头区域（所有行）
    for row_data in header_rows:
        new_ws.append(row_data)

    # 写入数据行
    for row_data in rows_data:
        new_ws.append(row_data)

    # 复制列宽
    for col_letter, width in column_widths.items():
        if width is not None:
            new_ws.column_dimensions[col_letter].width = width

    # 复制表头区域各行行高
    for i, height in enumerate(header_row_heights, start=1):
        if height is not None:
            new_ws.row_dimensions[i].height = height

    new_wb.save(out_path)
    new_wb.close()


def _sanitize_filename(filename: str) -> str:
    """清理文件名中的非法字符。"""
    filename = re.sub(r'[\\/:*?"<>|]', '_', filename)
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    return filename


def get_total_rows(filepath: str) -> int:
    """获取 Excel 文件的总行数。"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    max_row = ws.max_row or 0
    wb.close()
    return max_row


def split_excel(
    filepath: str,
    group_column: str,
    output_dir: str,
    filename_template: str = "{分组值}",
    progress_callback: Callable[[int, int, str], None] | None = None,
    header_row_num: int = 1,
) -> list[str]:
    """按指定列的唯一值拆分 Excel 文件，每组数据生成一个独立的 Excel 文件。

    参数:
        filepath: 源 Excel 文件路径
        group_column: 用于分组的列名
        output_dir: 输出目录
        filename_template: 文件名模板，支持 {分组值}、{序号} 变量
        progress_callback: 进度回调 (current, total, message)
        header_row_num: 表头所在行号（从 1 开始），默认 1。
            第 1 行到该行的所有内容都会作为表头区域保留到每个输出文件。
    返回:
        生成的文件路径列表
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    # 读取表头区域、列宽、行高
    header_rows, column_widths, header_row_heights = _read_header_area_and_widths(ws, header_row_num)

    # 用最后一行（即实际列名行）来定位分组列
    header_names = header_rows[-1]
    header_names_str = [str(h) if h is not None else "" for h in header_names]

    if group_column not in header_names_str:
        wb.close()
        raise ValueError(f"列 '{group_column}' 不存在，可用列: {', '.join(header_names_str)}")

    group_col_idx = header_names_str.index(group_column)

    # 按分组列收集数据行（从表头区域下一行开始）
    groups: dict[Any, list[list[Any]]] = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        group_value = row[group_col_idx] if group_col_idx < len(row) else None
        if group_value is None:
            group_value = ""
        group_value = str(group_value).strip()
        if group_value not in groups:
            groups[group_value] = []
        groups[group_value].append(list(row))

    total = len(groups)
    generated_files = []

    for i, (group_value, rows) in enumerate(groups.items(), start=1):
        if progress_callback:
            progress_callback(i, total, f"正在拆分第 {i}/{total} 组: {group_value}")

        # 生成文件名
        filename = filename_template
        filename = filename.replace("{分组值}", group_value)
        filename = filename.replace("{序号}", str(i).zfill(3))
        filename = _sanitize_filename(filename)

        out_path = os.path.join(output_dir, filename)

        new_wb = Workbook()
        new_ws = new_wb.active
        _save_workbook(
            new_wb, new_ws, ws.title, header_rows, rows,
            column_widths, header_row_heights, out_path
        )
        generated_files.append(out_path)

    wb.close()
    return generated_files


def split_excel_by_rows(
    filepath: str,
    rows_per_file: int,
    output_dir: str,
    filename_template: str = "{序号}",
    progress_callback: Callable[[int, int, str], None] | None = None,
    header_row_num: int = 1,
) -> list[str]:
    """按固定行数拆分 Excel 文件，每 N 行数据生成一个独立的 Excel 文件。

    参数:
        filepath: 源 Excel 文件路径
        rows_per_file: 每个文件的数据行数
        output_dir: 输出目录
        filename_template: 文件名模板，支持 {序号} 变量
        progress_callback: 进度回调 (current, total, message)
        header_row_num: 表头所在行号（从 1 开始），默认 1。
            第 1 行到该行的所有内容都会作为表头区域保留到每个输出文件。
    返回:
        生成的文件路径列表
    """
    if rows_per_file <= 0:
        raise ValueError("每份行数必须大于 0")

    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(filepath, data_only=False)
    ws = wb.active

    # 读取表头区域、列宽、行高
    header_rows, column_widths, header_row_heights = _read_header_area_and_widths(ws, header_row_num)

    # 收集所有数据行（从表头区域下一行开始）
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        all_rows.append(list(row))

    # 计算拆分份数
    total_files = (len(all_rows) + rows_per_file - 1) // rows_per_file
    if total_files == 0:
        total_files = 1  # 即使没有数据行，也至少生成一份带表头的文件

    generated_files = []

    for i in range(total_files):
        start = i * rows_per_file
        end = min(start + rows_per_file, len(all_rows))
        chunk = all_rows[start:end]

        if progress_callback:
            progress_callback(i + 1, total_files, f"正在拆分第 {i + 1}/{total_files} 份（{len(chunk)} 行）")

        # 生成文件名
        filename = filename_template
        filename = filename.replace("{序号}", str(i + 1).zfill(3))
        filename = _sanitize_filename(filename)

        out_path = os.path.join(output_dir, filename)

        new_wb = Workbook()
        new_ws = new_wb.active
        _save_workbook(
            new_wb, new_ws, ws.title, header_rows, chunk,
            column_widths, header_row_heights, out_path
        )
        generated_files.append(out_path)

    wb.close()
    return generated_files
